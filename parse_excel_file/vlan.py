"""Перед вами Excel файл с шаблонными таблицами.
   Задание:
   На странице IP-plan представлены таблицы по различным площадкам и доменам.
   Нужно собрать информацию вида: vlan name - vlan id - gateway - mask
   по площадке Москва, третий домен и вывести в консоль.
   Следует заметить, что домен делится на два сайта. Информацию можно вывести по одному."""
import openpyxl
import json
from pprint import pprint


AREA = 'Moscow'
DOMAIN = 'Domain3'


excel_file = openpyxl.load_workbook('ip_plan.xlsx', data_only=True)
ip_plan_sheet = excel_file['IP-plan']


def get_start_of_table(sheet, area, domain_number):
    for row in range(1, sheet.max_row+1):
        # Обработка пустых ячеек
        if sheet[f'A{row}'].value is None or sheet[f'B{row}'].value is None:
            continue
        # Обработка остальных
        elif area in sheet[f'A{row}'].value and domain_number in sheet[f'B{row}'].value:
            print(f'A{row}')
            return row + 5


def get_end_of_table(sheet, start_of_table):
    for row in range(start_of_table, sheet.max_row+1):
        if sheet[f'A{row}'].border.bottom.style == 'medium':
            return row + 1


def get_data_about_vlan(sheet, start_of_table, end_of_table):
    msk_dmn_3_vlan = []
    for row in range(start_of_table, end_of_table):
        if sheet[f'A{row}'].value is None or 'Supernet' in sheet[f'A{row}'].value:
            continue
        msk_dmn_3_vlan.append(
            {
                'name': sheet[f'A{row}'].value,
                'id': sheet[f'E{row}'].value,
                'gateway': sheet[f'G{row}'].value,
                'mask': sheet[f'C{row}'].value
            }
        )
    print(json.dumps(msk_dmn_3_vlan, indent=4, sort_keys=False))
    pprint(msk_dmn_3_vlan)


if __name__ == '__main__':
    start = get_start_of_table(ip_plan_sheet, AREA, DOMAIN)
    end = get_end_of_table(ip_plan_sheet, start)
    get_data_about_vlan(ip_plan_sheet, start, end)
