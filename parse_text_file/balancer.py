"""Перед вами текстовый файл с конфигурацией балансировщика.
   Задание:
   Вам нужно будет обработать файл, забрать оттуда определенные данные, создать файд excel c тремя страницами и
   заполнить на каждой странице таблицу.
   Пример excel файла с данными- example_for_balancer.
   1) Первый лист - Service. Название столбцов: Service, VIP, Port, Protocol, Pool.
   Вам следует найти блок данных, начинающийся с 'ltm virtual /Common/'. Значения, которые вам нужно запомнить,
   выделены для наглядности следующим образом: {{значение}}. Рядом написано к какому столбцу принадлежит значение.
   ltm virtual /Common/{{temp_atg}} - Service {
    creation-time 2021-02-20:19:12:12
    destination /Common/{{1.1.1.1}} - VIP:{{80}} - Port
    ip-protocol {{tcp}} - Protocol
    last-modified-time 2021-02-20:19:16:10
    mask 255.255.255.255
    persist {
        /Common/cookie {
            default yes
        }
    }
    pool /Common/{{atg-ps-prod_tcp80}} - Pool
    profiles {
        /Common/http { }
        /Common/tcp { }
    }
    source 0.0.0.0/0
    source-address-translation {
        pool /Common/SNAT_ATG
        type snat
    }
    translate-address enabled
    translate-port enabled
}
   2) Второй лист - Node. Название столбцов: Name, IP.
   Вам следует найти блок данных, начинающийся с 'ltm node /Common/'. Значения, которые вам нужно запомнить,
   выделены для наглядности следующим образом: {{значение}}. Рядом написано к какому столбцу принадлежит значение.
   ltm node /Common/{{atg-ps15-prod}} - Name {
    address {{10.220.4.146}} - IP
    monitor /Common/none
   3) Третий лист - Pool. Название столбцов: Pool Name, LB Method, Port, Member name, Member IP, Monitor.
   Вам следует найти блок данных, начинающийся с 'ltm pool /Common/'. Значения, которые вам нужно запомнить,
   выделены для наглядности следующим образом: {{значение}}. Рядом написано к какому столбцу принадлежит значение.
   ltm pool /Common/{{atg-ps-pilot_tcp80}} - Pool name {
    load-balancing-mode {{least-connections-member}} - LB Method (если данной записи нет, вывести в тиблицу RR)
    members {
        /Common/{{atg-ps16-prod}} - Member name:{{80}} - Port {
            address {{10.220.4.147}} - Member IP
        }
        /Common/{{atg-ps17-prod}}:{{80}} - Здесь все аналагично как в member выше {
            address {{10.220.4.148}}
        }
        /Common/{{atg-ps18-prod}}:{{80}} - И здесь {
            address {{10.220.4.149}}
        }
    }
    monitor /Common/{{tcp}} - Monitor
      """
from pprint import pprint
import openpyxl
import json


FILE = 'configs/balancer.conf1'


def read_file(name_file):
    with open(f'{name_file}.txt', 'r') as f:
        return f.readlines()


def parse_services(text):
    service = []
    for index, line in enumerate(text):
        if 'ltm virtual /Common/' in line:
            service.append(
                {
                    'service': line[20:-2]
                }
            )
            for next_index, next_line in enumerate(text[index+1::]):
                if 'ltm ' in next_line:
                    break
                elif '##ENDOF_CONFIG##' in next_line:
                    # pprint(pool, sort_dicts=False)
                    print(json.dumps(service, indent=4, sort_keys=False))
                    return service
                elif 'destination /Common/' in next_line:
                    service[len(service) - 1]['vip'] = next_line[24::].split(':')[0]
                    service[len(service) - 1]['port'] = next_line[24::].split(':')[1]
                elif 'ip-protocol' in next_line:
                    service[len(service) - 1]['protocol'] = next_line[16::]
                elif next_line.startswith('    pool /Common/'):
                    service[len(service) - 1]['pool'] = next_line[17::]
        elif '##ENDOF_CONFIG##' in line:
            # pprint(service, sort_dicts=False)
            print(json.dumps(service, indent=4, sort_keys=False))
            return service


def parse_nodes(text):
    node = []
    for index, line in enumerate(text):
        if 'ltm node /Common/' in line:
            node.append(
                {
                    'name': line[17:-2],
                    'ip': text[index+1][12::]
                }
            )
    # pprint(node, sort_dicts=False)
    print(json.dumps(node, indent=4, sort_keys=False))
    return node


def parse_pools(text):
    pool = []
    for pool_index, pool_line in enumerate(text):
        if 'ltm pool /Common/' in pool_line:
            pool.append(
                {
                    'name': pool_line[17:-2],
                }
            )
            pool[len(pool) - 1]['lb_method'] = 'RR'
            pool[len(pool) - 1]['members'] = []
            for index, line in enumerate(text[pool_index + 1::]):
                if 'ltm ' in line:
                    break
                elif '##ENDOF_CONFIG##' in line:
                    # pprint(pool, sort_dicts=False)
                    print(json.dumps(pool, indent=4, sort_keys=False))
                    return pool
                elif 'load-balancing-mode' in line:
                    pool[len(pool) - 1]['lb_method'] = line[23::]
                elif 'members' in line:
                    for member_index, member_line in enumerate(text[pool_index + 1::][index + 1::]):
                        if '}' in member_line and '}' in text[pool_index + 1::][index + 1::][member_index + 1]:
                            break
                        elif 'monitor /Common/' in member_line:
                            break
                        elif f'/Common/' in member_line:
                            pool[len(pool) - 1]['members'].append(
                                {
                                    'name': member_line[16:-2].split(':')[0],
                                    'port': member_line[16:-2].split(':')[1],
                                    'ip': text[pool_index + 1::][index + 1::][member_index + 1][20::]
                                }
                            )
                elif 'monitor' in line:
                    if 'and' in line:
                        monitors = line[20::].split('and')
                        monitors[1] = monitors[1][9:]
                        pool[len(pool) - 1]['monitor'] = ' '.join(monitors)
                    else:
                        pool[len(pool) - 1]['monitor'] = line[20::]
        elif '##ENDOF_CONFIG##' in pool_line:
            # pprint(pool, sort_dicts=False)
            print(json.dumps(pool, indent=4, sort_keys=False))
            return pool


def create_sheet_and_fill_in_the_column_names(excel_file, sheet_name, column_names):
    """Создание таблицы и заполнение названий колонок таблицы"""
    excel_file.create_sheet(sheet_name)
    sheet = excel_file[sheet_name]
    for column in range(1, len(column_names)+1):
        sheet.cell(row=1, column=column).value = column_names[column-1]


def push_data_into_service_sheet(excel_file, service_data):
    names_column_of_service_list = ('Service', 'VIP', 'Port', 'Protocol', 'Pool')
    create_sheet_and_fill_in_the_column_names(excel_file, 'Service', names_column_of_service_list)
    sheet = excel_file['Service']
    for index, service in enumerate(service_data):
        sheet[f'A{index + 2}'] = service['service']
        sheet[f'B{index + 2}'] = service['vip']
        sheet[f'C{index + 2}'] = service['port']
        sheet[f'D{index + 2}'] = service.get('protocol')
        sheet[f'E{index + 2}'] = service.get('pool')


def push_data_into_node_sheet(excel_file, node_data):
    names_column_of_node_list = ('Name', 'IP')
    create_sheet_and_fill_in_the_column_names(excel_file, 'Node', names_column_of_node_list)
    sheet = excel_file['Node']
    for index, node in enumerate(node_data):
        sheet[f'A{index + 2}'] = node['name']
        sheet[f'B{index + 2}'] = node['ip']


def push_data_into_pool_list(excel_file, pool_data):
    names_column_of_pool_list = ('Pool Name', 'LB Method', 'Port', 'Member name', 'Member IP', 'Monitor')
    create_sheet_and_fill_in_the_column_names(excel_file, 'Pool', names_column_of_pool_list)
    sheet = excel_file['Pool']
    cell_number = 2
    for pool in pool_data:
        sheet[f'A{cell_number}'] = pool['name']
        sheet[f'B{cell_number}'] = pool['lb_method']
        sheet[f'F{cell_number}'] = pool.get('monitor')
        for member in pool['members']:
            sheet[f'C{cell_number}'] = member['port']
            sheet[f'D{cell_number}'] = member['name']
            sheet[f'E{cell_number}'] = member['ip']
            cell_number += 1


def data_to_excel(excel_file, service_data, node_data, pool_data):
    # Заполнение Service List
    push_data_into_service_sheet(excel_file, service_data)
    # Заполнение Node List
    push_data_into_node_sheet(excel_file, node_data)
    # Заполнение Pool List
    push_data_into_pool_list(excel_file, pool_data)


if __name__ == '__main__':
    # Забираем из файла список строк и убиораем \n в конце каждой строки
    text_lines = [line.rstrip() for line in read_file(FILE)]
    services = parse_services(text_lines)
    nodes = parse_nodes(text_lines)
    pools = parse_pools(text_lines)
    file = openpyxl.Workbook()
    del file['Sheet']
    data_to_excel(file, services, nodes, pools)
    file.save(f'{FILE}.xlsx')
