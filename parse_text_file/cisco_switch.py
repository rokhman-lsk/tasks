"""Перед вами файл конфигурации коммутатора Cisco.
   Задание:
   1. Сопоставить номер vlan его символьному названию и вывести данное соответствие в консоль в виде словаря.
   2. Сопоставить названию интерфейса его ip-адрес, description и switchport mode
   (если хар-ки нет - поставить прочерк). Вывести соответствие в excel таблицу,
   где 1ая колонка интерфейс, 2ая - адрес и т.д.
   3. Заменить в строчке clock timezone - MSK на VRN"""
import re
from pprint import pprint
from openpyxl import Workbook


NAME_FILE = 'configs/cisco_switch.txt'


def read_file(name_file):
    with open(name_file, 'r') as file:
        return file.readlines()


def parse_vlan():
    vlan = {}
    text = read_file(NAME_FILE)
    for index, text_line in enumerate(text):
        if re.fullmatch(r'vlan \d+', text_line.strip()):
            vlan[f'{text_line[5:-1]}'] = text[index+1][6:-1] if 'name' in text[index+1] else '-'
    return vlan


def parse_interface():
    interface = {}
    text = read_file(NAME_FILE)
    for index, text_line in enumerate(text):
        if 'interface' in text_line:
            name_interface = text_line[10::].rstrip()
            interface[f'{name_interface}'] = 3*['-']
            while text[index+1][0] == ' ':
                next_line = text[index+1].rstrip()
                if 'description' in next_line:
                    interface[f'{name_interface}'][0] = next_line[13::]
                elif 'ip address' in next_line:
                    interface[f'{name_interface}'][1] = next_line[12::]
                elif 'switchport mode' in next_line:
                    interface[f'{name_interface}'][2] = next_line[17::]
                index += 1
    return interface


def interface_to_excel(interface):
    excel_file = Workbook()
    excel_file.active.title = 'Interfaces'
    sheet = excel_file['Interfaces']
    names_column = ('Interfaces', 'Description', 'IP', 'Mode')
    # Заполнение названий столбцов
    for column in range(1, 5):
        sheet.cell(row=1, column=column).value = names_column[column-1]
    # Заполнение данными
    index_row = 2
    for key, row_data in interface.items():
        sheet.cell(row=index_row, column=1).value = key
        for column in range(2, 5):
            sheet.cell(row=index_row, column=column).value = row_data[column-2]
        index_row += 1
    excel_file.save('interfaces.xlsx')


def change_tmz():
    text = read_file(NAME_FILE)
    for index, line in enumerate(text):
        if 'clock timezone MSK 3 0' in line:
            text[index] = text[index].replace('MSK', 'VRN')
            break
    with open(NAME_FILE, 'w') as file:
        file.writelines(text)


if __name__ == '__main__':
    pprint(parse_vlan())
    # print(parse_interface())
    # interface_to_excel(parse_interface())
    # change_tmz()
